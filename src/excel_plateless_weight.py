import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from calc import (ALL_NOMINAL_GAPS, HSDATA, PLATES_AND_SPACERS, WSDATA,
                  InputData, StepScreen)

THIN = Side(border_style='thin', color='000000')
COLOR = 'fce4d6'


def addborder(cell: Cell,
              top: Side | None = None,
              bottom: Side | None = None,
              left: Side | None = None,
              right: Side | None = None) -> None:
    old = cell.border
    cell.border = Border(top=top or old.top,
                         bottom=bottom or old.bottom,
                         left=left or old.left,
                         right=right or old.right)


def rectangle(cell_range: tuple[tuple[Cell]] | Cell) -> None:
    if isinstance(cell_range, Cell):
        addborder(cell_range, top=THIN, bottom=THIN, left=THIN, right=THIN)
        return
    for col, i in enumerate(cell_range[0]):
        addborder(i, top=THIN)
        addborder(cell_range[-1][col], bottom=THIN)
    for row in cell_range:
        addborder(row[0], left=THIN)
        addborder(row[-1], right=THIN)


def fillcolor(cell_range: tuple[tuple[Cell]] | Cell, color: str) -> None:
    if isinstance(cell_range, Cell):
        cell_range.fill = PatternFill('solid', fgColor=color)
        return
    for row in cell_range:
        for i in row:
            i.fill = PatternFill('solid', fgColor=color)


def main() -> None:
    book = Workbook()
    sheet = book.active

    row: int
    for hi, hs in enumerate(reversed(HSDATA)):
        row = 2 + hi
        sheet.cell(row=row, column=1, value=f'xx{hs:02d}')
        for wi, ws in enumerate(WSDATA):
            inp = InputData(ws=ws,
                            hs=hs,
                            gap=ALL_NOMINAL_GAPS[0],
                            depth=hs * 0.1,
                            plate_spacer=PLATES_AND_SPACERS[0],
                            steel_only=False)
            scr = StepScreen(inp, None)
            cell = sheet.cell(row=row, column=2 + wi)
            cell.value = round(scr.weight.plateless)
            if not scr.drive:
                fillcolor(cell, COLOR)

    row += 1
    for wi, ws in enumerate(WSDATA):
        cell = sheet.cell(row=row, column=2 + wi, value=f'{ws:02d}xx')
        cell.alignment = Alignment(horizontal='right')

    title = sheet.cell(row=1, column=1, value='Вага РСК без пластин (кг)')
    title.font = Font(b=True)
    title.alignment = Alignment(horizontal='center')
    sheet.merge_cells(start_row=1, start_column=1,
                      end_row=1, end_column=len(WSDATA) + 1)

    rectangle(sheet['A2:S11'])
    rectangle(sheet['A2:A11'])
    rectangle(sheet['A11:S11'])

    row += 2
    sheet.cell(row=row, column=1,
               value='Конструкція із пластиковим транспортером')
    sheet.cell(row=row + 1, column=1,
               value='Глибина каналу в мм дорівнює '
                     '(100 * Типорозмір за висотою)')
    sheet.cell(row=row + 2, column=1,
               value='Вага без привода, конструкція потребує адаптування')
    fillcolor(sheet['A15:I15'], COLOR)

    for i in range(1, len(WSDATA) + 2):
        sheet.column_dimensions[get_column_letter(i)].width = 6

    today = datetime.now().strftime('%Y-%m-%d')
    filename = f'RSK plateless weight ({today}).xlsx'
    book.save(filename)
    os.system(f'"{filename}"')


if __name__ == '__main__':
    main()
